"""
Teaching Hours Checker - Logic xử lý
Kiểm soát giờ dạy GV: đối chiếu Lịch kỳ FAP, Teaching Summaries, Phiếu chấm công ĐT
FPT Education QA Department
© 2026 YenLT31
"""

import pandas as pd
import numpy as np
import re
import unicodedata
import os
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 0. HÀM TRỢ GIÚP DỮ LIỆU
# ============================================================

def clean_text(val):
    """
    Chuẩn hóa chuỗi văn bản thông thường:
    - Chuẩn hóa Unicode về dựng sẵn (NFC)
    - Loại bỏ zero-width spaces và các ký tự ẩn
    - Collapse multiple spaces thành single space và strip
    """
    if pd.isna(val):
        return ""
    val_str = str(val)
    val_str = unicodedata.normalize('NFC', val_str)
    val_str = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', val_str)
    val_str = re.sub(r'\s+', ' ', val_str).strip()
    return val_str


def clean_account(val):
    """
    Chuẩn hóa AccountFE / Lecturer code viết tắt:
    - Chuẩn hóa Unicode dựng sẵn (NFC)
    - Loại bỏ hoàn toàn khoảng trắng để merge an toàn nhất
    - Chuyển thành chữ thường
    - Giữ lại các ký tự đặc biệt hợp lệ như dấu chấm (.) hay gạch dưới (_)
    """
    if pd.isna(val):
        return ""
    val_str = str(val)
    val_str = unicodedata.normalize('NFC', val_str)
    val_str = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', val_str)
    val_str = re.sub(r'\s+', '', val_str).strip().lower()
    return val_str


def normalize_header_key(val):
    """
    Convert Excel header text to a stable ASCII key for column detection.
    """
    val_str = clean_text(val).lower()
    val_str = unicodedata.normalize('NFD', val_str)
    val_str = ''.join(ch for ch in val_str if unicodedata.category(ch) != 'Mn')
    val_str = val_str.replace('đ', 'd')
    val_str = re.sub(r'[^a-z0-9.]+', ' ', val_str)
    return re.sub(r'\s+', ' ', val_str).strip()


def parse_date_robust(val):
    """
    Hàm phân tích ngày thông minh và chuẩn hóa về Timestamp (chỉ chứa ngày, normalize giờ về 00:00:00).
    Hỗ trợ:
      - Timestamp / datetime.datetime
      - Số serial date của Excel (float/int hoặc dạng chuỗi tương ứng, ví dụ "46097.0")
      - Các định dạng chuỗi ngày thông dụng: DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY, DD/MM/YY, v.v.
    """
    if pd.isna(val):
        return pd.NaT
    if isinstance(val, (datetime, pd.Timestamp)):
        return pd.to_datetime(val).normalize()
    
    val_str = str(val).strip()
    if not val_str or val_str.lower() == 'nan':
        return pd.NaT
    
    # 1. Nếu là số serial date của Excel dạng chuỗi (ví dụ: "46097" hoặc "46097.0")
    if re.match(r'^\d+(\.0+)?$', val_str):
        try:
            serial_date = int(float(val_str))
            # Excel epoch bắt đầu từ 1899-12-30 do lỗi năm nhuận 1900
            return pd.to_datetime(serial_date, unit='D', origin='1899-12-30').normalize()
        except Exception:
            pass

    year_first_match = re.match(
        r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?$',
        val_str
    )
    if year_first_match:
        year, month, day = [int(part) for part in year_first_match.groups()]
        try:
            return pd.Timestamp(year=year, month=month, day=day).normalize()
        except ValueError:
            pass

    numeric_date_match = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})$', val_str)
    if numeric_date_match:
        first, second, year = [int(part) for part in numeric_date_match.groups()]
        if year < 100:
            year += 2000
        if first > 12:
            day, month = first, second
        elif second > 12:
            month, day = first, second
        else:
            day, month = first, second
        try:
            return pd.Timestamp(year=year, month=month, day=day).normalize()
        except ValueError:
            pass
            
    # 2. Thử pd.to_datetime trước với dayfirst=True
    try:
        dt = pd.to_datetime(val_str, dayfirst=True, errors='coerce')
        if pd.notna(dt):
            return dt.normalize()
    except Exception:
        pass
        
    # 3. Thử parse theo các định dạng thủ công thông dụng
    formats = [
        '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', 
        '%d/%m/%y', '%d-%m-%y', '%m/%d/%Y', 
        '%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S',
        '%d/%b/%y', '%d/%b/%Y', '%d-%b-%y', '%d-%b-%Y'
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(val_str, fmt)
            return pd.Timestamp(dt).normalize()
        except ValueError:
            continue
            
    # 4. Thử pd.to_datetime tự do cuối cùng
    try:
        dt = pd.to_datetime(val_str, errors='coerce')
        if pd.notna(dt):
            return dt.normalize()
    except Exception:
        pass
        
    return pd.NaT


def infer_month_from_text(text):
    """
    Infer month number from a file/sheet name such as fap1.xlsx, FAP_01,
    thang 1, month-01 or T01. Return None when it is ambiguous.
    """
    if text is None:
        return None

    name = os.path.splitext(os.path.basename(str(text)))[0].lower()
    name = unicodedata.normalize('NFD', name)
    name = ''.join(ch for ch in name if unicodedata.category(ch) != 'Mn')
    name = name.replace('đ', 'd')

    patterns = [
        r'(?:^|[^a-z0-9])(?:fap|ts|teaching|summary|summaries)\s*[_\-\s]*(\d{1,2})(?:[^0-9]|$)',
        r'(?:^|[^a-z0-9])(?:thang|month|t)\s*[_\-\s]*(\d{1,2})(?:[^0-9]|$)',
    ]
    for pattern in patterns:
        match = re.search(pattern, name)
        if match:
            month = int(match.group(1))
            if 1 <= month <= 12:
                return month

    numbers = [int(n) for n in re.findall(r'\d{1,2}', name)]
    valid_months = [n for n in numbers if 1 <= n <= 12]
    return valid_months[0] if len(valid_months) == 1 else None


# ============================================================
# 1. ĐỌC VÀ CHUẨN HÓA DỮ LIỆU
# ============================================================

def read_lich_ky(files):
    """
    Đọc và gộp nhiều file Lịch kỳ FAP.
    Chuẩn hóa cột Date về datetime.
    """
    all_data = []
    for file in files:
        try:
            df = pd.read_excel(file, dtype=str)
            # Chuẩn hóa tên cột (strip whitespace)
            df.columns = df.columns.str.strip()
            all_data.append(df)
        except Exception as e:
            print(f"Lỗi đọc file lịch kỳ: {e}")
            continue
    
    if not all_data:
        return pd.DataFrame()
    
    df = pd.concat(all_data, ignore_index=True)
    
    # Chuẩn hóa cột Date sử dụng hàm parse_date_robust
    df['Date'] = df['Date'].apply(parse_date_robust)
    
    # Chuẩn hóa cột Lecturer: lowercase, strip, remove spaces
    df['Lecturer'] = df['Lecturer'].apply(clean_account)
    
    # Chuẩn hóa SubjectCode: strip, normalize
    df['SubjectCode'] = df['SubjectCode'].apply(clean_text)
    
    # Chuẩn hóa GroupName: strip, normalize
    df['GroupName'] = df['GroupName'].apply(clean_text)
    
    # Chuẩn hóa TypeSlot: strip, uppercase, normalize
    df['TypeSlot'] = df['TypeSlot'].apply(clean_text).str.upper()
    
    # Chuẩn hóa SlotTypeCode: strip, uppercase, normalize
    df['SlotTypeCode'] = df['SlotTypeCode'].apply(clean_text).str.upper()
    
    return df


def read_teaching_summaries(files):
    """
    Đọc nhiều file Teaching Summaries FAP.
    Mỗi file có thể có nhiều sheets (mỗi sheet là 1 tháng).
    Parse cấu trúc grouped (GV ở dòng riêng).
    Trả về DataFrame: Teacher, Group, Subject, AllPlan, WasNotTaken, FromDate, ToDate
    """
    all_data = []
    warnings = []
    
    for file in files:
        source_name = getattr(file, 'name', str(file))
        source_month_from_file = infer_month_from_text(source_name)
        try:
            # Đọc tất cả sheets trong file
            xl = pd.ExcelFile(file)
            
            for sheet_name in xl.sheet_names:
                source_month = source_month_from_file or infer_month_from_text(sheet_name)
                df = pd.read_excel(file, sheet_name=sheet_name, header=None)
                
                # Bỏ qua sheet trống
                if df.empty or len(df) < 5:
                    continue
                
                # Tìm From Date và To Date
                from_date = None
                to_date = None
                
                for idx in range(min(30, len(df))):
                    row = df.iloc[idx]
                    row_str = [str(v).strip() for v in row.values if pd.notna(v)]
                    row_str_lower = [s.lower() for s in row_str]
                    if 'from date' in row_str_lower:
                        # Dòng tiếp theo là giá trị
                        if idx + 1 < len(df):
                            next_row = df.iloc[idx + 1]
                            from_date = parse_date_robust(next_row.iloc[0])
                            to_date = parse_date_robust(next_row.iloc[1])
                        break
                
                # Tìm header row (Teacher, Group, Subject, All Plan, WasNot Taken)
                header_row_idx = None
                for idx in range(min(30, len(df))):
                    row = df.iloc[idx]
                    row_str = [str(v).strip().lower() for v in row.values if pd.notna(v)]
                    # Phải chứa đồng thời ít nhất 3 từ khóa cốt lõi để tránh nhận nhầm ô dữ liệu thông thường
                    has_teacher = any('teacher' in s for s in row_str)
                    has_group = any('group' in s or 'class' in s for s in row_str)
                    has_subject = any('subject' in s for s in row_str)
                    
                    if has_teacher and has_group and has_subject:
                        header_row_idx = idx
                        break
                
                if header_row_idx is None:
                    warnings.append(
                        f"Teaching Summaries sheet '{sheet_name}': khong tim thay header chua Teacher, Group/Class va Subject trong 30 dong dau."
                    )
                    continue
                
                # Đọc lại với header đúng
                df_data = df.iloc[header_row_idx + 1:].copy()
                headers = df.iloc[header_row_idx].values
                
                # Xử lý duplicate column names
                seen = {}
                clean_headers = []
                for h in headers:
                    h_str = str(h).strip() if pd.notna(h) else 'unnamed'
                    if h_str in seen:
                        seen[h_str] += 1
                        clean_headers.append(f"{h_str}_{seen[h_str]}")
                    else:
                        seen[h_str] = 0
                        clean_headers.append(h_str)
                
                df_data.columns = clean_headers
                
                # Tìm cột tương ứng (lấy cột đầu tiên match)
                teacher_col = None
                group_col = None
                subject_col = None
                allplan_col = None
                wasnot_col = None
                
                for col in df_data.columns:
                    col_lower = str(col).strip().lower()
                    if 'teacher' in col_lower and teacher_col is None:
                        teacher_col = col
                    elif 'group' in col_lower and group_col is None:
                        group_col = col
                    elif 'subject' in col_lower and subject_col is None:
                        subject_col = col
                    elif 'all plan' in col_lower and allplan_col is None:
                        allplan_col = col
                    elif ('wasnot' in col_lower or 'was not' in col_lower) and wasnot_col is None:
                        wasnot_col = col
                
                # Parse grouped structure
                current_teacher = None
                records = []
                
                for idx, row in df_data.iterrows():
                    teacher_val = str(row[teacher_col]).strip() if teacher_col and pd.notna(row[teacher_col]) else ''
                    group_val = str(row[group_col]).strip() if group_col and pd.notna(row[group_col]) else ''
                    subject_val = str(row[subject_col]).strip() if subject_col and pd.notna(row[subject_col]) else ''
                    allplan_val = row[allplan_col] if allplan_col else None
                    wasnot_val = str(row[wasnot_col]).strip() if wasnot_col and pd.notna(row[wasnot_col]) else ''
                    
                    # Loại bỏ giá trị 'nan'
                    if teacher_val.lower() == 'nan':
                        teacher_val = ''
                    if group_val.lower() == 'nan':
                        group_val = ''
                    if subject_val.lower() == 'nan':
                        subject_val = ''
                    if wasnot_val.lower() == 'nan':
                        wasnot_val = ''
                    
                    # Nếu có Teacher và không có Group → đây là dòng tên GV
                    if teacher_val and not group_val:
                        current_teacher = clean_account(teacher_val)
                    # Nếu có Group và Subject → đây là dòng dữ liệu
                    elif group_val and subject_val:
                        clean_grp = clean_text(group_val)
                        clean_sub = clean_text(subject_val)
                        # Parse số WasNotTaken từ text
                        wasnot_count = 0
                        if wasnot_val:
                            match = re.search(r'(\d+)\s*slot', wasnot_val, re.IGNORECASE)
                            if match:
                                wasnot_count = int(match.group(1))
                        
                        # Parse AllPlan
                        try:
                            allplan_count = int(float(allplan_val)) if pd.notna(allplan_val) else 0
                        except (ValueError, TypeError):
                            allplan_count = 0
                        
                        records.append({
                            'Teacher': current_teacher if current_teacher else '',
                            'Group': clean_grp,
                            'Subject': clean_sub,
                            'AllPlan': allplan_count,
                            'WasNotTaken': wasnot_count,
                            'WasNotTakenRaw': wasnot_val,
                            'FromDate': from_date,
                            'ToDate': to_date,
                            'SourceFile': os.path.basename(str(source_name)),
                            'SourceSheet': sheet_name,
                            'SourceMonth': source_month
                        })
                
                if records:
                    all_data.extend(records)
        
        except Exception as e:
            warnings.append(f"Loi doc file Teaching Summaries: {e}")
            continue
    
    if not all_data:
        if warnings:
            raise ValueError("Khong doc duoc Teaching Summaries. " + " ".join(warnings))
        return pd.DataFrame()
    
    result = pd.DataFrame(all_data)
    return result


def read_cham_cong(files):
    """
    Đọc nhiều file Phiếu chấm công ĐT (sheet "Gio day cua thang").
    Parse thông tin tháng, khoảng ngày, giờ dạy.
    Xử lý merged cells.
    """
    all_data = []
    warnings = []
    
    for file in files:
        try:
            # Tìm sheet phù hợp
            xl = pd.ExcelFile(file)
            target_sheet = None
            
            for s in xl.sheet_names:
                s_lower = s.lower().replace(' ', '').replace('_', '')
                if 'giodaycuathang' in s_lower or 'giờdạycủatháng' in s_lower:
                    target_sheet = s
                    break
                elif 'gioday' in s_lower or 'giờdạy' in s_lower:
                    target_sheet = s
                    break
            
            if target_sheet is None:
                target_sheet = xl.sheet_names[0]
            
            df = pd.read_excel(file, sheet_name=target_sheet, header=None)
            
            # Parse thông tin tháng và khoảng ngày
            thang = None
            from_date = None
            to_date = None
            
            for idx in range(min(30, len(df))):
                row = df.iloc[idx]
                for col_idx, val in enumerate(row.values):
                    if pd.notna(val):
                        val_str = str(val).strip()
                        if 'NĂM' in val_str.upper() or 'NAM' in val_str.upper():
                            for next_col in range(col_idx + 1, len(row.values)):
                                next_val = row.values[next_col]
                                if pd.notna(next_val):
                                    next_str = str(next_val).strip()
                                    thang_match = re.search(r'(\d{1,2}/\d{4})', next_str)
                                    if thang_match:
                                        t_val = thang_match.group(1)
                                        if '/' in t_val:
                                            m, y = t_val.split('/')
                                            thang = f"{int(m):02d}/{y}"
                                    
                                    date_match = re.search(
                                        r'[Tt]ừ\s*(\d{1,2}/\d{1,2}/\d{4})\s*[đđ]ến\s*(\d{1,2}/\d{1,2}/\d{4})',
                                        next_str
                                    )
                                    if date_match:
                                        from_date = parse_date_robust(date_match.group(1))
                                        to_date = parse_date_robust(date_match.group(2))
                                    break
                            
            if thang is None:
                warnings.append(
                    f"Cham cong sheet '{target_sheet}': khong tim thay thong tin thang trong 30 dong dau."
                )
                continue
            
            # Tìm vị trí các cột quan trọng
            if from_date is not None and to_date is not None and pd.notna(from_date) and pd.notna(to_date):
                thang_month, thang_year = [int(part) for part in thang.split('/')]
                is_valid_cross_year_january = (
                    thang_month == 1 and
                    from_date.month == 12 and
                    from_date.year == thang_year - 1 and
                    to_date.year == thang_year
                )
                if not is_valid_cross_year_january and from_date.year != thang_year:
                    from_date = from_date.replace(year=thang_year)

            id_col = None
            ho_ten_col = None
            don_vi_col = None
            bo_mon_col = None
            doi_tuong_col = None
            hs1_col = None
            hs13_col = None
            header_end_row = None
            
            for idx in range(min(30, len(df))):
                row = df.iloc[idx]
                header_keys = [normalize_header_key(val) for val in row.values]
                has_id = any(key == 'id' for key in header_keys)
                has_name = any('ho ten' in key or key == 'hoten' for key in header_keys)
                
                if not (has_id and has_name):
                    continue
                
                header_end_row = idx
                for col_idx, key in enumerate(header_keys):
                    if key == 'id':
                        id_col = col_idx
                    elif 'ho ten' in key or key == 'hoten':
                        ho_ten_col = col_idx
                    elif key == 'don vi' or 'don vi' in key:
                        don_vi_col = col_idx
                    elif key == 'bo mon' or 'bo mon' in key:
                        bo_mon_col = col_idx
                    elif key == 'doi tuong' or 'doi tuong' in key:
                        doi_tuong_col = col_idx
                    elif 'he so 1.3' in key:
                        hs13_col = col_idx
                    elif 'he so 1' in key and '1.3' not in key:
                        hs1_col = col_idx
                
                if don_vi_col is not None:
                    if bo_mon_col is None and don_vi_col + 1 < len(header_keys):
                        bo_mon_col = don_vi_col + 1
                if bo_mon_col is not None and bo_mon_col + 1 < len(header_keys):
                    doi_tuong_col = bo_mon_col + 1
                break

            # Cac cot gio day co the nam o header nhieu tang, khong cung dong voi ID/Ho ten.
            for idx in range(min(30, len(df))):
                row = df.iloc[idx]
                for col_idx, val in enumerate(row.values):
                    key = normalize_header_key(val)
                    if 'he so 1.3' in key:
                        hs13_col = col_idx
                    elif 'he so 1' in key and '1.3' not in key:
                        hs1_col = col_idx
            
            if id_col is None or hs1_col is None:
                warnings.append(
                    f"Cham cong sheet '{target_sheet}': khong tim thay header ID va He so 1 trong 30 dong dau."
                )
                continue
            
            # Tìm dòng bắt đầu dữ liệu (sau header, có ID là số)
            data_start = (header_end_row + 1) if header_end_row else 0
            for idx in range(data_start, min(data_start + 10, len(df))):
                val = df.iloc[idx, id_col]
                if pd.notna(val):
                    val_clean = re.sub(r'\.0$', '', str(val).strip())
                    if re.match(r'^\d+$', val_clean):
                        data_start = idx
                        break

            # Đọc từng dòng dữ liệu
            for idx in range(data_start, len(df)):
                row = df.iloc[idx]
                
                id_val = str(row.iloc[id_col]).strip() if pd.notna(row.iloc[id_col]) else ''
                # Loại bỏ .0 nếu pandas đọc số thành float
                id_val = re.sub(r'\.0$', '', id_val)
                
                # Bỏ qua dòng trống hoặc không phải ID số
                if not id_val or id_val == 'nan' or not re.match(r'^\d+$', id_val):
                    continue
                
                ho_ten = clean_text(row.iloc[ho_ten_col]) if ho_ten_col is not None and pd.notna(row.iloc[ho_ten_col]) else ''
                don_vi = clean_text(row.iloc[don_vi_col]) if don_vi_col is not None and pd.notna(row.iloc[don_vi_col]) else ''
                bo_mon = clean_text(row.iloc[bo_mon_col]) if bo_mon_col is not None and pd.notna(row.iloc[bo_mon_col]) else ''
                doi_tuong = clean_text(row.iloc[doi_tuong_col]) if doi_tuong_col is not None and pd.notna(row.iloc[doi_tuong_col]) else ''
                
                # Loại bỏ 'nan'
                if ho_ten.lower() == 'nan':
                    ho_ten = ''
                if don_vi.lower() == 'nan':
                    don_vi = ''
                if bo_mon.lower() == 'nan':
                    bo_mon = ''
                if doi_tuong.lower() == 'nan':
                    doi_tuong = ''
                
                # Giờ dạy
                hs1 = 0
                hs13 = 0
                try:
                    hs1 = float(row.iloc[hs1_col]) if pd.notna(row.iloc[hs1_col]) else 0
                except (ValueError, TypeError):
                    hs1 = 0
                try:
                    if hs13_col is not None:
                        hs13 = float(row.iloc[hs13_col]) if pd.notna(row.iloc[hs13_col]) else 0
                except (ValueError, TypeError):
                    hs13 = 0
                
                gio_day_dt = hs1 + hs13
                
                all_data.append({
                    'Thang': thang,
                    'FromDate': from_date,
                    'ToDate': to_date,
                    'ID': id_val,
                    'HoTen': ho_ten,
                    'DonVi': don_vi,
                    'BoMon': bo_mon,
                    'DoiTuong': doi_tuong,
                    'HeSo1': hs1,
                    'HeSo13': hs13,
                    'GioDayDT': gio_day_dt
                })
        
        except Exception as e:
            warnings.append(f"Loi doc file cham cong: {e}")
            continue
    
    if not all_data:
        if warnings:
            raise ValueError("Khong doc duoc file cham cong. " + " ".join(warnings))
        return pd.DataFrame()
    
    result = pd.DataFrame(all_data)
    return result


def read_danh_sach_gv(file):
    """
    Đọc file Danh sách GV (mapping).
    Trả về DataFrame: ID, TeacherFullname, AccountFE, Major, Note
    """
    df = pd.read_excel(file, dtype=str)
    df.columns = df.columns.str.strip()
    
    # Chuẩn hóa tên cột
    col_mapping = {}
    drop_cols = []
    for col in df.columns:
        col_lower = col.strip().lower()
        if col_lower in ['no.', 'no', 'stt']:
            drop_cols.append(col)
        elif col_lower == 'id':
            col_mapping[col] = 'ID'
        elif 'fullname' in col_lower or 'full name' in col_lower or "teacher" in col_lower:
            col_mapping[col] = 'TeacherFullname'
        elif 'accountfe' in col_lower or 'account' in col_lower:
            col_mapping[col] = 'AccountFE'
        elif 'major' in col_lower:
            col_mapping[col] = 'Major'
        elif 'note' in col_lower:
            col_mapping[col] = 'Note'
    
    # Xóa cột STT
    if drop_cols:
        df = df.drop(columns=drop_cols)
    
    df = df.rename(columns=col_mapping)
    
    # Chuẩn hóa
    if 'ID' in df.columns:
        df['ID'] = df['ID'].astype(str).str.strip()
        # Loại bỏ .0 nếu pandas đọc thành float
        df['ID'] = df['ID'].str.replace(r'\.0$', '', regex=True)
        # Loại bỏ dòng ID trống hoặc nan
        df = df[df['ID'].notna() & (df['ID'] != '') & (df['ID'] != 'nan')]
    
    if 'AccountFE' in df.columns:
        df['AccountFE'] = df['AccountFE'].apply(clean_account)
        # Loại bỏ dòng AccountFE trống
        df = df[df['AccountFE'].notna() & (df['AccountFE'] != '') & (df['AccountFE'] != 'nan')]
    
    return df


# ============================================================
# 2. LOGIC TÍNH TOÁN
# ============================================================

def get_type_slot_mapping(df_lich_ky):
    """
    Tạo mapping SubjectCode → TypeSlot từ lịch kỳ.
    Mỗi SubjectCode chỉ có 1 TypeSlot duy nhất.
    Bỏ qua SlotTypeCode = 'G' khi tạo mapping.
    """
    df_filtered = df_lich_ky[df_lich_ky['SlotTypeCode'] != 'G'].copy()
    
    mapping = {}
    for _, row in df_filtered.drop_duplicates(subset=['SubjectCode']).iterrows():
        subject = str(row['SubjectCode']).strip()
        type_slot = str(row['TypeSlot']).strip().upper()
        mapping[subject] = type_slot
    
    return mapping


def get_hours_per_slot(type_slot):
    """
    Trả về số giờ cho mỗi buổi dựa vào TypeSlot.
    NEW SLOT = 2.25h, OLD SLOT = 1.5h
    """
    type_slot_upper = str(type_slot).upper()
    if 'NEW' in type_slot_upper:
        return 2.25
    elif 'OLD' in type_slot_upper:
        return 1.5
    else:
        return 2.25  # Default NEW SLOT


def calculate_gio_lich_ky(df_lich_ky, from_date, to_date):
    """
    Tính tổng giờ Lịch kỳ mỗi GV trong khoảng ngày.
    Lọc bỏ SlotTypeCode = 'G'.
    Chưa trừ WasNot Taken.
    """
    # Lọc theo khoảng ngày
    mask = (df_lich_ky['Date'] >= from_date) & (df_lich_ky['Date'] <= to_date)
    df_filtered = df_lich_ky[mask].copy()
    
    # Lọc bỏ SlotTypeCode = 'G'
    df_filtered = df_filtered[df_filtered['SlotTypeCode'] != 'G']
    
    if df_filtered.empty:
        return pd.DataFrame(columns=['AccountFE', 'GioLichKy'])
    
    # Tính giờ cho mỗi buổi
    df_filtered['Hours'] = df_filtered['TypeSlot'].apply(get_hours_per_slot)
    
    # Tổng giờ theo GV (Lecturer = AccountFE lowercase)
    result = df_filtered.groupby('Lecturer')['Hours'].sum().reset_index()
    result.columns = ['AccountFE', 'GioLichKy']
    
    return result


def calculate_slots_lich_ky(df_lich_ky, from_date, to_date):
    """
    Tinh tong slot Lich ky moi GV trong khoang ngay.
    Loc bo SlotTypeCode = 'G'. Moi dong lich ky tuong ung 1 slot.
    """
    mask = (df_lich_ky['Date'] >= from_date) & (df_lich_ky['Date'] <= to_date)
    df_filtered = df_lich_ky[mask].copy()
    df_filtered = df_filtered[df_filtered['SlotTypeCode'] != 'G']

    if df_filtered.empty:
        return pd.DataFrame(columns=['AccountFE', 'SlotsLichKy'])

    result = df_filtered.groupby('Lecturer').size().reset_index(name='SlotsLichKy')
    result.columns = ['AccountFE', 'SlotsLichKy']

    return result


def calculate_slots_fap(df_teaching_summaries, from_date, to_date):
    """
    Tính tổng giờ FAP mỗi GV.
    Giờ FAP = Σ từng lớp: (AllPlan - WasNotTaken) × giờ/buổi (theo TypeSlot của SubjectCode)
    Match TypeSlot qua SubjectCode từ lịch kỳ.
    """
    if df_teaching_summaries.empty:
        return pd.DataFrame(columns=['AccountFE', 'SlotsFAP'])
    
    from_date = parse_date_robust(from_date)
    to_date = parse_date_robust(to_date)
    
    df_teaching_summaries = df_teaching_summaries.copy()
    df_teaching_summaries['FromDate'] = df_teaching_summaries['FromDate'].apply(parse_date_robust)
    df_teaching_summaries['ToDate'] = df_teaching_summaries['ToDate'].apply(parse_date_robust)
    
    mask = (df_teaching_summaries['FromDate'] == from_date) & (df_teaching_summaries['ToDate'] == to_date)
    df_filtered = df_teaching_summaries[mask].copy()
    
    if df_filtered.empty:
        overlap_mask = (
            df_teaching_summaries['FromDate'].notna() &
            df_teaching_summaries['ToDate'].notna() &
            (df_teaching_summaries['FromDate'] <= to_date) &
            (df_teaching_summaries['ToDate'] >= from_date)
        )
        df_filtered = df_teaching_summaries[overlap_mask].copy()

    # Some FAP exports contain the From Date / To Date labels but no actual
    # date values. In that case, match monthly files by SourceMonth inferred
    # from file/sheet names, for example fap1.xlsx -> month 1.
    if df_filtered.empty and 'SourceMonth' in df_teaching_summaries.columns:
        target_month = None
        if pd.notna(to_date):
            target_month = int(to_date.month)
        elif pd.notna(from_date):
            target_month = int(from_date.month)

        if target_month is not None:
            source_month = pd.to_numeric(df_teaching_summaries['SourceMonth'], errors='coerce')
            df_filtered = df_teaching_summaries[source_month == target_month].copy()
    
    if df_filtered.empty:
        return pd.DataFrame(columns=['AccountFE', 'SlotsFAP'])
    
    # Lấy mapping SubjectCode → TypeSlot từ lịch kỳ (bỏ qua G)
    
    # Tính giờ từng dòng
    def calc_slots(row):
        actual_slots = row['AllPlan'] - row['WasNotTaken']
        if actual_slots < 0:
            actual_slots = 0
        return actual_slots
    
    df_filtered = df_filtered.copy()
    df_filtered['Slots'] = df_filtered.apply(calc_slots, axis=1)
    
    # Tổng theo GV
    result = df_filtered.groupby('Teacher')['Slots'].sum().reset_index()
    result.columns = ['AccountFE', 'SlotsFAP']
    
    return result


def calculate_gio_dt(df_cham_cong, thang):
    """
    Lấy tổng giờ ĐT mỗi GV từ phiếu chấm công cho tháng cụ thể.
    Giờ ĐT = Hệ số 1 + Hệ số 1.3
    """
    df_filtered = df_cham_cong[df_cham_cong['Thang'] == thang].copy()
    
    if df_filtered.empty:
        return pd.DataFrame(columns=['ID', 'HoTen', 'DonVi', 'BoMon', 'DoiTuong', 'GioDayDT'])
    
    result = df_filtered[['ID', 'HoTen', 'DonVi', 'BoMon', 'DoiTuong', 'GioDayDT']].copy()
    return result


# ============================================================
# 3. ĐỐI SÁNH VÀ TẠO BÁO CÁO
# ============================================================

def doi_sanh_gio_day(df_lich_ky, df_teaching_summaries, df_cham_cong, df_gv_mapping):
    """
    Đối sánh giờ dạy: Lịch kỳ vs FAP vs ĐT.
    Trả về DataFrame kết quả gộp nhiều tháng.
    """
    results = []
    
    # Chuẩn hóa ID (loại bỏ .0, strip)
    df_gv_mapping = df_gv_mapping.copy()
    if 'ID' in df_gv_mapping.columns:
        df_gv_mapping['ID'] = df_gv_mapping['ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    
    df_cham_cong = df_cham_cong.copy()
    df_cham_cong['ID'] = df_cham_cong['ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    
    # Lấy danh sách tháng từ chấm công
    months = df_cham_cong[['Thang', 'FromDate', 'ToDate']].drop_duplicates()
    
    for _, month_row in months.iterrows():
        thang = month_row['Thang']
        from_date = month_row['FromDate']
        to_date = month_row['ToDate']
        
        if pd.isna(from_date) or pd.isna(to_date):
            continue
        
        # Tính giờ Lịch kỳ
        gio_lich_ky = calculate_gio_lich_ky(df_lich_ky, from_date, to_date)
        slots_lich_ky = calculate_slots_lich_ky(df_lich_ky, from_date, to_date)
        
        # Tính giờ FAP
        slots_fap = calculate_slots_fap(df_teaching_summaries, from_date, to_date)
        
        # Lấy giờ ĐT
        gio_dt = calculate_gio_dt(df_cham_cong, thang)
        
        # === Tạo bảng master từ Danh sách GV ===
        if 'ID' in df_gv_mapping.columns and 'AccountFE' in df_gv_mapping.columns:
            master = df_gv_mapping[['ID', 'AccountFE']].copy()
            master = master.dropna(subset=['AccountFE'])
            master['AccountFE'] = master['AccountFE'].apply(clean_account)
            # Loại bỏ AccountFE = 'nan'
            master = master[master['AccountFE'] != 'nan']
        else:
            continue
        
        # Merge giờ ĐT vào master qua ID (chỉ lấy GV có trong chấm công)
        master = master.merge(
            gio_dt[['ID', 'HoTen', 'DonVi', 'BoMon', 'DoiTuong', 'GioDayDT']],
            on='ID',
            how='inner'
        )
        
        # Merge giờ Lịch kỳ vào master qua AccountFE
        master = master.merge(gio_lich_ky, on='AccountFE', how='left')
        
        # Merge giờ FAP vào master qua AccountFE
        master = master.merge(slots_lich_ky, on='AccountFE', how='left')
        master = master.merge(slots_fap, on='AccountFE', how='left')
        
        # Fill NaN = 0 cho cột giờ
        master['GioLichKy'] = master['GioLichKy'].fillna(0)
        master['SlotsLichKy'] = master['SlotsLichKy'].fillna(0)
        master['SlotsFAP'] = master['SlotsFAP'].fillna(0)
        master['GioDayDT'] = master['GioDayDT'].fillna(0)
        
        # Tính chênh lệch
        master['ChenhLech_LichKy_DT'] = master['GioLichKy'] - master['GioDayDT']
        master['CheckSlots_LichKy_FAP'] = master['SlotsLichKy'] == master['SlotsFAP']
        
        # Kết quả TRUE/FALSE
        master['KetQua'] = (
            (master['ChenhLech_LichKy_DT'] == 0) &
            (master['CheckSlots_LichKy_FAP'])
        )
        
        master['Thang'] = thang
        results.append(master)
    
    if not results:
        return pd.DataFrame()
    
    final = pd.concat(results, ignore_index=True)
    
    # Sắp xếp cột
    cols_order = ['Thang', 'ID', 'HoTen', 'AccountFE', 'DonVi', 'BoMon', 'DoiTuong',
                  'GioLichKy', 'GioDayDT', 'ChenhLech_LichKy_DT',
                  'SlotsLichKy', 'SlotsFAP', 'CheckSlots_LichKy_FAP', 'KetQua']
    existing_cols = [c for c in cols_order if c in final.columns]
    final = final[existing_cols]
    
    # Sắp xếp theo Tháng, HoTen
    final = final.sort_values(['Thang', 'HoTen']).reset_index(drop=True)
    
    return final


def get_wasnot_taken_detail(df_teaching_summaries):
    """
    Trả về chi tiết các buổi WasNot Taken.
    """
    if df_teaching_summaries.empty:
        return pd.DataFrame()
    
    df_filtered = df_teaching_summaries[df_teaching_summaries['WasNotTaken'] > 0].copy()
    
    if df_filtered.empty:
        return pd.DataFrame()
    
    df_filtered['Thang'] = ''
    if 'SourceMonth' in df_filtered.columns:
        source_month = pd.to_numeric(df_filtered['SourceMonth'], errors='coerce')
        df_filtered.loc[source_month.notna(), 'Thang'] = (
            source_month[source_month.notna()].astype(int).astype(str).str.zfill(2) + '/2026'
        )

    if 'FromDate' in df_filtered.columns:
        from_month = df_filtered['FromDate'].apply(parse_date_robust)
        has_from_month = df_filtered['Thang'].eq('') & from_month.notna()
        df_filtered.loc[has_from_month, 'Thang'] = from_month[has_from_month].dt.strftime('%m/%Y')

    base_cols = ['Thang', 'Teacher', 'Group', 'Subject', 'WasNotTaken', 'WasNotTakenRaw']
    optional_cols = [col for col in ['SourceFile', 'SourceSheet', 'FromDate', 'ToDate'] if col in df_filtered.columns]
    result = df_filtered[base_cols + optional_cols].copy()
    result = result.rename(columns={
        'Teacher': 'AccountFE',
        'WasNotTaken': 'SoBuoiNghi',
        'WasNotTakenRaw': 'ChiTiet'
    })

    group_cols = ['Thang', 'AccountFE', 'Group', 'Subject']
    agg_spec = {
        'SoBuoiNghi': 'sum',
        'ChiTiet': lambda values: ' | '.join(sorted({str(v).strip() for v in values if str(v).strip()}))
    }
    for col in ['SourceFile', 'SourceSheet', 'FromDate', 'ToDate']:
        if col in result.columns:
            agg_spec[col] = 'first'

    result = result.groupby(group_cols, dropna=False, as_index=False).agg(agg_spec)
    ordered_cols = ['Thang', 'AccountFE', 'Group', 'Subject', 'SoBuoiNghi', 'ChiTiet']
    ordered_cols += [col for col in ['SourceFile', 'SourceSheet', 'FromDate', 'ToDate'] if col in result.columns]
    return result[ordered_cols].sort_values(['Thang', 'AccountFE', 'Group', 'Subject']).reset_index(drop=True)


# ============================================================
# 4. TÍNH GIỜ DẠY CƠ HỮU
# ============================================================

def is_co_huu(doi_tuong):
    """
    Kiểm tra đối tượng có phải cơ hữu không.
    Cơ hữu: CBNV, CBQL, CH, CHdn, CHNN1, GVNCV, GVQL hoặc bắt đầu bằng 'CH'
    """
    if not doi_tuong or str(doi_tuong).strip() == '' or str(doi_tuong).strip().lower() == 'nan':
        return False
    
    doi_tuong = str(doi_tuong).strip().upper()
    co_huu_list = ['CBNV', 'CBQL', 'CH', 'CHDN', 'CHNN1', 'GVNCV', 'GVQL']
    
    if doi_tuong in co_huu_list:
        return True
    if doi_tuong.startswith('CH'):
        return True
    
    return False


def calculate_gio_co_huu(df_lich_ky_full, df_cham_cong, df_gv_mapping):
    """
    Tính giờ dạy cơ hữu HK.
    Nguồn giờ: Lịch kỳ FAP toàn kỳ (KHÔNG bỏ SlotTypeCode = 'G')
    Đối tượng: Từ chấm công ĐT
    """
    # Tính giờ mỗi GV từ lịch kỳ (KHÔNG lọc SlotTypeCode G)
    df_calc = df_lich_ky_full.copy()
    df_calc['Hours'] = df_calc['TypeSlot'].apply(get_hours_per_slot)
    
    gio_per_gv = df_calc.groupby('Lecturer')['Hours'].sum().reset_index()
    gio_per_gv.columns = ['AccountFE', 'TongGio']
    
    # Chuẩn hóa ID
    df_gv_mapping = df_gv_mapping.copy()
    if 'ID' in df_gv_mapping.columns:
        df_gv_mapping['ID'] = df_gv_mapping['ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    
    df_cham_cong = df_cham_cong.copy()
    df_cham_cong['ID'] = df_cham_cong['ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    
    # Lấy đối tượng từ chấm công (lấy unique theo ID)
    doi_tuong_mapping = df_cham_cong[['ID', 'HoTen', 'DoiTuong', 'BoMon', 'DonVi']].drop_duplicates(subset=['ID'])
    
    # Merge với danh sách GV để có AccountFE
    if 'ID' in df_gv_mapping.columns and 'AccountFE' in df_gv_mapping.columns:
        doi_tuong_mapping = doi_tuong_mapping.merge(
            df_gv_mapping[['ID', 'AccountFE']],
            on='ID',
            how='left'
        )
    
    if 'AccountFE' not in doi_tuong_mapping.columns:
        return pd.DataFrame(), 0, 0
    
    # Chuẩn hóa AccountFE
    doi_tuong_mapping['AccountFE'] = doi_tuong_mapping['AccountFE'].apply(clean_account)
    doi_tuong_mapping = doi_tuong_mapping[
        doi_tuong_mapping['AccountFE'].notna() &
        (doi_tuong_mapping['AccountFE'] != '') &
        (doi_tuong_mapping['AccountFE'] != 'nan')
    ]
    
    # Merge giờ với đối tượng
    result = doi_tuong_mapping.merge(
        gio_per_gv,
        on='AccountFE',
        how='left'
    )
    result['TongGio'] = result['TongGio'].fillna(0)
    
    # Xác định cơ hữu
    
    # Tính tổng
    tong_gio_all = result['TongGio'].sum()
    tong_gio_co_huu = result[result['DoiTuong'].apply(is_co_huu) == True]['TongGio'].sum()
    
    cols_order = ['AccountFE', 'HoTen', 'DoiTuong', 'BoMon', 'DonVi', 'TongGio']
    result = result[cols_order].sort_values(['AccountFE']).reset_index(drop=True)
    
    return result, tong_gio_co_huu, tong_gio_all


def build_lich_ky_kiem_tra(df_lich_ky_full, df_cham_cong, df_gv_mapping):
    """
    Tao sheet lich ky kiem tra: them Thang theo khoang cham cong va
    thong tin GV link qua ID -> AccountFE.
    """
    if df_lich_ky_full.empty:
        return pd.DataFrame()

    df_check = df_lich_ky_full.copy()
    df_check['Hours'] = df_check['TypeSlot'].apply(get_hours_per_slot)
    df_check['Thang'] = ''
    df_check['FromDate'] = pd.NaT
    df_check['ToDate'] = pd.NaT

    periods = df_cham_cong[['Thang', 'FromDate', 'ToDate']].drop_duplicates().copy()
    periods['FromDate'] = periods['FromDate'].apply(parse_date_robust)
    periods['ToDate'] = periods['ToDate'].apply(parse_date_robust)

    for _, period in periods.iterrows():
        from_date = period['FromDate']
        to_date = period['ToDate']
        if pd.isna(from_date) or pd.isna(to_date):
            continue

        thang_label = f"{int(to_date.month)}/{int(to_date.year)}"
        mask = (df_check['Date'] >= from_date) & (df_check['Date'] <= to_date)
        df_check.loc[mask, 'Thang'] = thang_label
        df_check.loc[mask, 'FromDate'] = from_date
        df_check.loc[mask, 'ToDate'] = to_date

    df_gv_mapping = df_gv_mapping.copy()
    if 'ID' in df_gv_mapping.columns:
        df_gv_mapping['ID'] = df_gv_mapping['ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    if 'AccountFE' in df_gv_mapping.columns:
        df_gv_mapping['AccountFE'] = df_gv_mapping['AccountFE'].apply(clean_account)

    df_cham_cong = df_cham_cong.copy()
    df_cham_cong['ID'] = df_cham_cong['ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    gv_info = df_cham_cong[['ID', 'HoTen', 'DoiTuong', 'BoMon', 'DonVi']].drop_duplicates(subset=['ID'])

    if 'ID' in df_gv_mapping.columns and 'AccountFE' in df_gv_mapping.columns:
        gv_info = gv_info.merge(df_gv_mapping[['ID', 'AccountFE']], on='ID', how='left')
        gv_info['AccountFE'] = gv_info['AccountFE'].apply(clean_account)
        df_check = df_check.merge(
            gv_info[['AccountFE', 'ID', 'HoTen', 'DoiTuong', 'BoMon', 'DonVi']],
            left_on='Lecturer',
            right_on='AccountFE',
            how='left'
        )
    else:
        df_check['AccountFE'] = df_check['Lecturer']

    if 'AccountFE' in df_check.columns:
        df_check['AccountFE'] = df_check['AccountFE'].fillna(df_check['Lecturer'])

    cols_order = [
        'Thang', 'AccountFE', 'ID', 'HoTen',
        'DoiTuong', 'BoMon', 'DonVi', 'SubjectCode', 'GroupName',
        'TypeSlot', 'SlotTypeCode', 'Hours'
    ]
    existing_cols = [col for col in cols_order if col in df_check.columns]
    return df_check.sort_values(['Date', 'AccountFE'])[existing_cols].reset_index(drop=True)


# ============================================================
# 5. XUẤT FILE EXCEL
# ============================================================

def export_to_excel(df_doi_sanh, df_wasnot_taken, df_co_huu=None, df_lich_ky_kiem_tra=None):
    """
    Xuất kết quả ra file Excel với nhiều sheet và định dạng chuyên nghiệp.
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Đối sánh chi tiết
        if df_doi_sanh is not None and not df_doi_sanh.empty:
            df_export = df_doi_sanh.copy()
            df_export.to_excel(writer, sheet_name='Đối sánh giờ dạy', index=False)
        
        # Sheet 2: WasNot Taken
        if df_wasnot_taken is not None and not df_wasnot_taken.empty:
            df_wasnot_taken.to_excel(writer, sheet_name='WasNot Taken', index=False)
        
        # Sheet 3: Giờ cơ hữu (nếu có)
        if df_co_huu is not None and not df_co_huu.empty:
            df_co_huu.to_excel(writer, sheet_name='Giờ dạy cơ hữu', index=False)

        if df_lich_ky_kiem_tra is not None and not df_lich_ky_kiem_tra.empty:
            df_lich_ky_kiem_tra.to_excel(writer, sheet_name='Lịch kỳ kiểm tra', index=False)
            
    # Đọc lại bằng openpyxl để apply styles chuyên nghiệp
    output.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(output)
    
    # Định nghĩa styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Navy Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    mismatch_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Đỏ nhạt
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Kích hoạt gridlines
        ws.views.sheetView[0].showGridLines = True
        
        # Lấy số dòng và cột
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 1. Định dạng Header (Dòng 1)
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = cell_border
        
        # 2. Định dạng nội dung
        for row in range(2, max_row + 1):
            # Check xem dòng này có bị lệch kết quả đối sánh không (đối với sheet 'Đối sánh giờ dạy')
            is_mismatch = False
            if sheet_name == 'Đối sánh giờ dạy':
                # Tìm cột có tiêu đề 'KetQua'
                ket_qua_col_idx = None
                for col in range(1, max_col + 1):
                    if ws.cell(row=1, column=col).value == 'KetQua':
                        ket_qua_col_idx = col
                        break
                if ket_qua_col_idx:
                    val = ws.cell(row=row, column=ket_qua_col_idx).value
                    # Nếu KetQua là False hoặc chuỗi 'False'
                    if val is False or str(val).strip().upper() == 'FALSE':
                        is_mismatch = True
            
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = cell_border
                
                # Highlight nếu bị lệch
                if is_mismatch:
                    cell.fill = mismatch_fill
                
                # Căn lề theo kiểu dữ liệu
                val = cell.value
                col_name = ws.cell(row=1, column=col).value
                
                # Các cột dạng mã/ID/ngày/boolean/Tháng -> Căn giữa
                if col_name in ['Thang', 'ID', 'AccountFE', 'KetQua', 'CheckSlots_LichKy_FAP', 'FromDate', 'ToDate', 'Group', 'Subject', 'LaCoHuu', 'SoBuoiNghi']:
                    cell.alignment = align_center
                elif col_name and 'Slots' in col_name:
                    cell.alignment = align_right
                    cell.number_format = '#,##0'
                    try:
                        if val is not None:
                            cell.value = int(float(val))
                    except ValueError:
                        pass
                # Các cột dạng số/giờ -> Căn phải và format số
                elif col_name in ['Hours', 'TongGio']:
                    cell.alignment = align_right
                    cell.number_format = '#,##0.00'
                    try:
                        if val is not None:
                            cell.value = float(val)
                    except ValueError:
                        pass
                elif isinstance(val, (int, float)) or (col_name and ('Gio' in col_name or 'HeSo' in col_name or 'ChenhLech' in col_name)):
                    cell.alignment = align_right
                    cell.number_format = '#,##0.00'
                    try:
                        if val is not None:
                            cell.value = float(val)
                    except ValueError:
                        pass
                else:
                    cell.alignment = align_left
        
        # 3. Tự động co giãn cột (Auto-fit Columns)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                # Nếu là tiêu đề thì cộng thêm một khoảng đệm lớn hơn
                if cell.row == 1:
                    max_len = max(max_len, len(val_str) + 4)
                else:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    # Lưu lại
    output_styled = BytesIO()
    wb.save(output_styled)
    output_styled.seek(0)
    return output_styled
